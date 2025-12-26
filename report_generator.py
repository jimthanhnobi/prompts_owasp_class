"""
Report Generator - Generates comprehensive test reports
Following LLM_Test_Design_Framework_Template.xlsx format
"""
import json
import os
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config import TestConfig, OWASP_RISKS, CLASS_PRINCIPLES
from models import TestRunResult, TestSummary, PassFailStatus


class ReportGenerator:
    """Generates comprehensive test reports following template format"""
    
    def __init__(self, config: TestConfig):
        self.config = config
        Path(self.config.results_dir).mkdir(parents=True, exist_ok=True)
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _calc_security_pass_rate(self, results: List[TestRunResult]) -> float:
        """Calculate security test pass rate"""
        security_tests = [r for r in results if r.test_case_id.startswith("SEC_")]
        if not security_tests:
            return 100.0
        passed = len([r for r in security_tests if r.pass_fail == PassFailStatus.PASS])
        return (passed / len(security_tests)) * 100
    
    def generate_excel_report(
        self,
        results: List[TestRunResult],
        summary: TestSummary,
        test_cases: List[Dict] = None,
        output_path: Optional[str] = None
    ) -> str:
        """Generate comprehensive Excel report following template format"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if not output_path:
            output_path = Path(self.config.results_dir) / f"test_report_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Sheet 1: 00_Summary
        ws_overview = wb.active
        ws_overview.title = "00_Summary"
        self._create_framework_overview(ws_overview, summary)
        
        # Group results by category
        results_by_category, tc_map = self._group_results_by_category(results, test_cases or [])
        
        # Sheet 2: 01_Test_Results_Functional
        if results_by_category.get("Functional"):
            ws_func = wb.create_sheet("01_Test_Results_Functional")
            self._create_category_test_results(ws_func, results_by_category["Functional"], "Functional", tc_map)
        
        # Sheet 3: 02_Test_Results_OWASP
        if results_by_category.get("Security"):
            ws_owasp = wb.create_sheet("02_Test_Results_OWASP")
            self._create_category_test_results(ws_owasp, results_by_category["Security"], "Security", tc_map)
        
        # Sheet 4: 03_Test_Results_CLASS
        if results_by_category.get("C-L-A-S-S"):
            ws_class = wb.create_sheet("03_Test_Results_CLASS")
            self._create_category_test_results(ws_class, results_by_category["C-L-A-S-S"], "C-L-A-S-S", tc_map)
        
        # Sheet 5: 04_Test_Results_CLASS_Design
        if results_by_category.get("CLASS_Design"):
            ws_design = wb.create_sheet("04_Test_Results_CLASS_Design")
            self._create_category_test_results(ws_design, results_by_category["CLASS_Design"], "CLASS_Design", tc_map)
        
        # Sheet 6: 05_Metrics_C_L_A_S_S
        ws_metrics = wb.create_sheet("05_Metrics_C_L_A_S_S")
        self._create_metrics_classs(ws_metrics, results, summary)
        
        # Sheet 7: 06_OWASP_Coverage_Matrix
        ws_owasp_cov = wb.create_sheet("06_OWASP_Coverage_Matrix")
        self._create_owasp_coverage(ws_owasp_cov, results)
        
        # Sheet 8: 07_CLASS_Checklist
        ws_class_checklist = wb.create_sheet("07_CLASS_Checklist")
        self._create_class_checklist(ws_class_checklist, results)
        
        # Sheet 9: 08_Thresholds_Comparison
        ws_thresholds = wb.create_sheet("08_Thresholds_Comparison")
        self._create_thresholds_comparison(ws_thresholds, results, summary)
        
        # Sheet 10: 09_Workload_Analysis
        ws_workload = wb.create_sheet("09_Workload_Analysis")
        self._create_workload_analysis(ws_workload, results, summary)
        
        wb.save(output_path)
        return str(output_path)
    
    def _group_results_by_category(self, results: List[TestRunResult], test_cases: List[Dict]) -> Tuple[Dict[str, List], Dict]:
        """Group test results by category and return test case map"""
        # Create test case lookup with full info
        tc_map = {}
        for tc in test_cases:
            if isinstance(tc, dict):
                tc_id = tc.get("Test_Case_ID", "")
                category = tc.get("Category", "")
            else:
                tc_id = getattr(tc, 'test_case_id', '')
                category = getattr(tc, 'category', '')
            if tc_id:
                tc_map[tc_id] = {"category": category, "test_case": tc}
        
        # Group results
        grouped = {
            "Functional": [],
            "Security": [],
            "C-L-A-S-S": [],
            "CLASS_Design": []
        }
        
        for result in results:
            tc_info = tc_map.get(result.test_case_id, {})
            category = tc_info.get("category", "")
            
            if category in grouped:
                grouped[category].append(result)
            else:
                # Try to infer from test case ID
                if result.test_case_id.startswith("TC_"):
                    grouped["Functional"].append(result)
                elif result.test_case_id.startswith("SEC_"):
                    grouped["Security"].append(result)
                elif result.test_case_id.startswith("CLASSS_"):
                    grouped["C-L-A-S-S"].append(result)
                elif result.test_case_id.startswith("CLASS_"):
                    grouped["CLASS_Design"].append(result)
        
        return grouped, tc_map
    
    def _create_category_test_results(self, ws, results: List[TestRunResult], category: str, tc_map: Dict):
        """Create test results sheet for a specific category"""
        headers = [
            "Test_Case_ID", "Feature_Area", "Description_VN", "Priority",
            "User_Message_Input", "Expected_Response", "Actual_Response",
            "Pass_Fail", "Accuracy_%", "Latency_ms", "Cost_VND",
            "Security_Status", "Stability_Status"
        ]
        
        # Add category-specific columns
        if category == "Security":
            headers.append("OWASP_Risks")
            headers.append("OWASP_Result")
        elif category in ["C-L-A-S-S", "CLASS_Design"]:
            headers.append("CLASS_Dimensions")
            if category == "CLASS_Design":
                headers.append("CLASS_Principles")
        
        headers.append("Notes")
        
        self._apply_header_style(ws, 1, headers)
        
        row_idx = 2
        for result in results:
            col = 1
            
            # Get test case info
            tc_info = tc_map.get(result.test_case_id, {})
            tc = tc_info.get("test_case", {})
            
            # Extract test case info
            if isinstance(tc, dict):
                feature = tc.get("Feature_Area", "")
                desc = tc.get("Description_VN", "")
                priority = tc.get("Priority", "Medium")
                user_msg = tc.get("User_Message_Input", "")
                expected_resp = tc.get("Expected_Bot_Response", "")
            else:
                feature = getattr(tc, 'feature_area', '')
                desc = getattr(tc, 'description_vn', '')
                priority = getattr(tc, 'priority', 'Medium')
                user_msg = getattr(tc, 'user_message_input', '')
                expected_resp = getattr(tc, 'expected_bot_response', '')
            
            # Basic info
            ws.cell(row=row_idx, column=col, value=result.test_case_id).border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=feature).border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=desc).border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=priority).border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=user_msg).border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=expected_resp).border = self.thin_border
            col += 1
            
            # Actual response
            actual_resp = result.actual_bot_response or ""
            if len(actual_resp) > 150:
                actual_resp = actual_resp[:150] + "..."
            ws.cell(row=row_idx, column=col, value=actual_resp).border = self.thin_border
            col += 1
            
            # Pass/Fail with color
            status_cell = ws.cell(row=row_idx, column=col, value=result.pass_fail.value)
            status_cell.border = self.thin_border
            if result.pass_fail == PassFailStatus.PASS:
                status_cell.fill = self.pass_fill
            elif result.pass_fail == PassFailStatus.FAIL:
                status_cell.fill = self.fail_fill
            elif result.pass_fail == PassFailStatus.PARTIAL:
                status_cell.fill = self.partial_fill
            col += 1
            
            ws.cell(row=row_idx, column=col, value=f"{result.accuracy_score_percent:.1f}").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=result.measured_latency_ms).border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=f"{result.measured_cost_vnd:.0f}").border = self.thin_border
            col += 1
            
            # Security status
            sec_cell = ws.cell(row=row_idx, column=col, value=result.security_observation.value)
            sec_cell.border = self.thin_border
            if result.security_observation.value != "OK":
                sec_cell.fill = self.fail_fill
            col += 1
            
            # Stability status
            stab_cell = ws.cell(row=row_idx, column=col, value=result.stability_observation.value)
            stab_cell.border = self.thin_border
            if result.stability_observation.value not in ["OK", "High_latency"]:
                stab_cell.fill = self.fail_fill
            elif result.stability_observation.value == "High_latency":
                stab_cell.fill = self.partial_fill
            col += 1
            
            # Category-specific columns
            if category == "Security":
                # OWASP Risks
                owasp_risks = ""
                if isinstance(tc, dict):
                    risks = tc.get("Target_OWASP_Risks", [])
                else:
                    risks = getattr(tc, 'target_owasp_risks', [])
                owasp_risks = ",".join(risks) if isinstance(risks, list) else str(risks or "")
                ws.cell(row=row_idx, column=col, value=owasp_risks).border = self.thin_border
                col += 1
                
                # OWASP Result
                owasp_result = ""
                if result.owasp_check:
                    parts = []
                    for risk_id, status in result.owasp_check.items():
                        if status == "OK":
                            parts.append(f"{risk_id}:✓")
                        else:
                            parts.append(f"{risk_id}:✗ {status}")
                    owasp_result = "; ".join(parts)
                
                owasp_cell = ws.cell(row=row_idx, column=col, value=owasp_result)
                owasp_cell.border = self.thin_border
                if "✗" in owasp_result:
                    owasp_cell.fill = self.fail_fill
                elif owasp_result:
                    owasp_cell.fill = self.pass_fill
                col += 1
            
            elif category in ["C-L-A-S-S", "CLASS_Design"]:
                # CLASS Dimensions
                class_dims = ""
                if isinstance(tc, dict):
                    dims = tc.get("Target_Dimensions_CLASSS", [])
                else:
                    dims = getattr(tc, 'target_dimensions_classs', [])
                class_dims = ",".join(dims) if isinstance(dims, list) else str(dims or "")
                ws.cell(row=row_idx, column=col, value=class_dims).border = self.thin_border
                col += 1
                
                if category == "CLASS_Design":
                    # CLASS Principles
                    class_principles = ""
                    if isinstance(tc, dict):
                        principles = tc.get("Target_CLASS_Principles", [])
                    else:
                        principles = getattr(tc, 'target_class_principles', [])
                    class_principles = ",".join(principles) if isinstance(principles, list) else str(principles or "")
                    ws.cell(row=row_idx, column=col, value=class_principles).border = self.thin_border
                    col += 1
            
            # Notes
            ws.cell(row=row_idx, column=col, value=result.notes or "").border = self.thin_border
            
            row_idx += 1
        
        # Set column widths dynamically based on number of columns
        num_cols = len(headers)
        base_widths = [15, 15, 40, 10, 40, 35, 40, 10, 10, 12, 12, 15, 15]  # Base columns
        if category == "Security":
            base_widths.extend([20, 30])  # OWASP columns
        elif category == "C-L-A-S-S":
            base_widths.append(20)  # CLASS Dimensions
        elif category == "CLASS_Design":
            base_widths.extend([20, 25])  # CLASS Dimensions + Principles
        base_widths.append(50)  # Notes
        
        for i, w in enumerate(base_widths[:num_cols]):
            ws.column_dimensions[chr(65 + i)].width = w
    
    def _apply_header_style(self, ws, row, headers):
        """Apply header style to a row"""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(wrap_text=True)
    
    def _create_framework_overview(self, ws, summary: TestSummary):
        """Create 00_Framework_Overview sheet"""
        headers = ["Section", "Item", "Value_Example", "Notes"]
        self._apply_header_style(ws, 1, headers)
        
        # Metadata section
        data = [
            ("Metadata", "System_Name", "MoneyCare Chatbot AI", "Chatbot ghi nhận giao dịch từ message người dùng"),
            ("Metadata", "Version", "v1.0", "Phiên bản dùng trong đợt test"),
            ("Metadata", "Test_Period_From", summary.start_time.strftime("%Y-%m-%d") if summary.start_time else datetime.now().strftime("%Y-%m-%d"), "Ngày bắt đầu test"),
            ("Metadata", "Test_Period_To", summary.end_time.strftime("%Y-%m-%d") if summary.end_time else datetime.now().strftime("%Y-%m-%d"), "Ngày kết thúc test"),
            ("Metadata", "LLM_Provider", "OpenAI", ""),
            ("Metadata", "LLM_Model", self.config.llm_model, ""),
            ("Metadata", "Owner", "Test Team", "PM / Tech Lead phụ trách"),
            ("Metadata", "Environment", self.config.environment, "Dev / Staging / Prod-simulated"),
            ("", "", "", ""),
            ("Framework_Mapping", "C - Cost", "Chi phí / 1k requests, tổng cost / ngày", "Chi tiết định nghĩa ở sheet 02_Metrics_C_L_A_S_S"),
            ("Framework_Mapping", "L - Latency", "Avg / P95 latency (ms)", ""),
            ("Framework_Mapping", "A - Accuracy", "Tỷ lệ parse đúng amount, category, date", ""),
            ("Framework_Mapping", "S - Scalability", "Max QPS, concurrent users", ""),
            ("Framework_Mapping", "S - Stability", "Error rate, uptime", ""),
            ("", "", "", ""),
            ("Test_Summary", "Total_Tests", str(summary.total_tests), ""),
            ("Test_Summary", "Passed", str(summary.passed), ""),
            ("Test_Summary", "Failed", str(summary.failed), ""),
            ("Test_Summary", "Partial", str(summary.partial), ""),
            ("Test_Summary", "Pass_Rate_%", f"{summary.pass_rate():.1f}", ""),
            ("Test_Summary", "Avg_Latency_ms", f"{summary.avg_latency_ms:.0f}", ""),
            ("Test_Summary", "Avg_Accuracy_%", f"{summary.avg_accuracy:.1f}", ""),
            ("Test_Summary", "Security_Issues", str(summary.security_issues), ""),
            ("Test_Summary", "Stability_Issues", str(summary.stability_issues), ""),
        ]
        
        for row_idx, row_data in enumerate(data, start=2):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50

    def _create_merged_test_results(self, ws, test_cases: List, results: List[TestRunResult]):
        """Create 01_Test_Results sheet - merged Test_Cases + Test_Run_Log + OWASP"""
        headers = [
            # Test Case Info
            "Test_Case_ID", "Feature_Area", "Description_VN", "Priority",
            # Input/Expected
            "User_Message_Input", "Expected_Response",
            # Actual Results
            "Actual_Response",
            # Evaluation
            "Pass_Fail", "Accuracy_%", "Latency_ms",
            # Security & Stability
            "Security_Status", "Stability_Status",
            # OWASP Result (gộp từ sheet 04)
            "OWASP_Check", "OWASP_Result",
            # Notes
            "Notes"
        ]
        self._apply_header_style(ws, 1, headers)
        
        # Create test case lookup
        tc_map = {}
        for tc in test_cases:
            if hasattr(tc, 'test_case_id'):
                tc_map[tc.test_case_id] = tc
            else:
                tc_map[tc.get("Test_Case_ID", "")] = tc
        
        for row_idx, result in enumerate(results, start=2):
            tc = tc_map.get(result.test_case_id, {})
            
            # Extract test case info
            if hasattr(tc, 'test_case_id'):
                feature = tc.feature_area
                desc = tc.description_vn
                priority = tc.priority
                user_msg = tc.user_message_input
                expected_resp = tc.expected_bot_response
                risks = tc.target_owasp_risks
            else:
                feature = tc.get("Feature_Area", "")
                desc = tc.get("Description_VN", "")
                priority = tc.get("Priority", "Medium")
                user_msg = tc.get("User_Message_Input", "")
                expected_resp = tc.get("Expected_Bot_Response", "")
                risks = tc.get("Target_OWASP_Risks", [])
            
            # Format actual response (truncated)
            actual_resp = result.actual_bot_response or ""
            if len(actual_resp) > 150:
                actual_resp = actual_resp[:150] + "..."
            
            # Format OWASP check results
            owasp_risks_str = ",".join(risks) if isinstance(risks, list) else str(risks or "")
            owasp_result_str = ""
            if result.owasp_check:
                owasp_parts = []
                for risk_id, status in result.owasp_check.items():
                    if status == "OK":
                        owasp_parts.append(f"{risk_id}:✓")
                    else:
                        owasp_parts.append(f"{risk_id}:✗ {status}")
                owasp_result_str = "; ".join(owasp_parts)
            
            # Write row
            ws.cell(row=row_idx, column=1, value=result.test_case_id).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=feature).border = self.thin_border
            ws.cell(row=row_idx, column=3, value=desc).border = self.thin_border
            ws.cell(row=row_idx, column=4, value=priority).border = self.thin_border
            ws.cell(row=row_idx, column=5, value=user_msg).border = self.thin_border
            ws.cell(row=row_idx, column=6, value=expected_resp).border = self.thin_border
            ws.cell(row=row_idx, column=7, value=actual_resp).border = self.thin_border
            
            # Pass/Fail with color
            status_cell = ws.cell(row=row_idx, column=8, value=result.pass_fail.value)
            status_cell.border = self.thin_border
            if result.pass_fail == PassFailStatus.PASS:
                status_cell.fill = self.pass_fill
            elif result.pass_fail == PassFailStatus.FAIL:
                status_cell.fill = self.fail_fill
            elif result.pass_fail == PassFailStatus.PARTIAL:
                status_cell.fill = self.partial_fill
            
            ws.cell(row=row_idx, column=9, value=f"{result.accuracy_score_percent:.1f}").border = self.thin_border
            ws.cell(row=row_idx, column=10, value=result.measured_latency_ms).border = self.thin_border
            
            # Security status with color
            sec_cell = ws.cell(row=row_idx, column=11, value=result.security_observation.value)
            sec_cell.border = self.thin_border
            if result.security_observation.value != "OK":
                sec_cell.fill = self.fail_fill
            
            # Stability status with color
            stab_cell = ws.cell(row=row_idx, column=12, value=result.stability_observation.value)
            stab_cell.border = self.thin_border
            if result.stability_observation.value not in ["OK", "High_latency"]:
                stab_cell.fill = self.fail_fill
            elif result.stability_observation.value == "High_latency":
                stab_cell.fill = self.partial_fill
            
            # OWASP Check (target risks)
            ws.cell(row=row_idx, column=13, value=owasp_risks_str).border = self.thin_border
            
            # OWASP Result with color
            owasp_cell = ws.cell(row=row_idx, column=14, value=owasp_result_str)
            owasp_cell.border = self.thin_border
            if owasp_result_str:
                if "✗" in owasp_result_str:
                    owasp_cell.fill = self.fail_fill
                else:
                    owasp_cell.fill = self.pass_fill
            
            ws.cell(row=row_idx, column=15, value=result.notes or "").border = self.thin_border
        
        # Set column widths
        widths = [12, 12, 35, 8, 45, 35, 45, 8, 10, 10, 15, 12, 15, 35, 35]
        for i, w in enumerate(widths):
            if i < 15:
                ws.column_dimensions[chr(65 + i)].width = w
    
    def _create_metrics_classs(self, ws, results: List[TestRunResult], summary: TestSummary):
        """Create 03_Metrics_C_L_A_S_S sheet"""
        headers = [
            "Dimension", "Metric_ID", "Metric_Name", "Definition", "Unit",
            "Data_Source", "Acceptable_Threshold", "Alert_Threshold",
            "Actual_Value", "Status", "Related_OWASP_Risks", "Related_CLASS_Principles"
        ]
        self._apply_header_style(ws, 1, headers)
        
        # Calculate metrics
        latencies = [r.measured_latency_ms for r in results if r.measured_latency_ms > 0]
        avg_latency = sum(latencies) / len(latencies) if latencies else 0
        p95_latency = sorted(latencies)[int(len(latencies) * 0.95)] if len(latencies) > 1 else avg_latency
        max_latency = max(latencies) if latencies else 0
        min_latency = min(latencies) if latencies else 0
        
        accuracies = [r.accuracy_score_percent for r in results if r.accuracy_score_percent > 0]
        avg_accuracy = sum(accuracies) / len(accuracies) if accuracies else 0
        
        error_count = len([r for r in results if r.pass_fail == PassFailStatus.ERROR])
        error_rate = (error_count / len(results) * 100) if results else 0
        
        timeout_count = len([r for r in results if r.stability_observation.value == "Timeout"])
        timeout_rate = (timeout_count / len(results) * 100) if results else 0
        
        # Cost estimation for GPT-4o-mini
        # Pricing: Input $0.15/1M tokens, Output $0.60/1M tokens
        # Estimate: ~500 input tokens + ~200 output tokens per request
        est_input_tokens_per_req = 500
        est_output_tokens_per_req = 200
        input_cost_per_1k = (est_input_tokens_per_req / 1000) * self.config.input_token_rate * 1000  # USD per 1k requests
        output_cost_per_1k = (est_output_tokens_per_req / 1000) * self.config.output_token_rate * 1000  # USD per 1k requests
        total_cost_usd_per_1k = input_cost_per_1k + output_cost_per_1k
        total_cost_vnd_per_1k = total_cost_usd_per_1k * self.config.usd_to_vnd_rate
        
        # Actual cost if we have token data, otherwise use estimate
        if summary.total_cost_vnd > 0:
            actual_cost_per_1k = summary.total_cost_vnd * 1000 / max(len(results), 1)
        else:
            actual_cost_per_1k = total_cost_vnd_per_1k
        
        # Daily cost estimate (assuming 1000 requests/day)
        daily_requests_estimate = 1000
        daily_cost_vnd = actual_cost_per_1k * (daily_requests_estimate / 1000)
        
        metrics_data = [
            ("C", "C1", "Cost_per_1k_requests", f"Chi phí LLM ước tính cho 1000 requests ({self.config.llm_model})", "VND/1k req", "Estimate: ~500 input + 200 output tokens/req", "≤ 100,000", "> 150,000", f"{actual_cost_per_1k:,.0f}", "OK" if actual_cost_per_1k <= 100000 else "Warning", "LLM10", ""),
            ("C", "C2", "Cost_per_request", "Chi phí trung bình mỗi request", "VND/req", "Calculated", "≤ 100", "> 150", f"{actual_cost_per_1k/1000:,.1f}", "OK" if actual_cost_per_1k/1000 <= 100 else "Warning", "", ""),
            ("C", "C3", "Daily_LLM_Cost_Est", f"Ước tính chi phí/ngày ({daily_requests_estimate:,} requests)", "VND/ngày", "Estimate", "≤ 100,000", "> 200,000", f"{daily_cost_vnd:,.0f}", "OK" if daily_cost_vnd <= 100000 else "Warning", "LLM10", ""),
            ("C", "C4", "Monthly_LLM_Cost_Est", "Ước tính chi phí/tháng (30 ngày)", "VND/tháng", "Estimate", "≤ 3,000,000", "> 6,000,000", f"{daily_cost_vnd * 30:,.0f}", "Info", "", ""),
            ("L", "L1", "Avg_Latency", "Thời gian phản hồi trung bình", "ms", "Test results", f"< {self.config.latency_warning_ms:,}", f"> {self.config.latency_critical_ms:,}", f"{avg_latency:,.0f}", "OK" if avg_latency < self.config.latency_warning_ms else ("Critical" if avg_latency > self.config.latency_critical_ms else "Warning"), "", ""),
            ("L", "L2", "P95_Latency", "P95 latency (95% requests nhanh hơn)", "ms", "Test results", f"< {self.config.latency_critical_ms:,}", "> 10,000", f"{p95_latency:,.0f}", "OK" if p95_latency < self.config.latency_critical_ms else ("Critical" if p95_latency > 10000 else "Warning"), "", ""),
            ("L", "L3", "Min_Latency", "Latency thấp nhất", "ms", "Test results", "", "", f"{min_latency:,.0f}", "Info", "", ""),
            ("L", "L4", "Max_Latency", "Latency cao nhất", "ms", "Test results", "< 15,000", "> 20,000", f"{max_latency:,.0f}", "OK" if max_latency < 15000 else "Warning", "", ""),
            ("L", "L5", "High_Latency_Rate", f"Tỷ lệ request > {self.config.latency_critical_ms}ms", "%", "Test results", "< 15%", "> 25%", f"{timeout_rate:.1f}", "OK" if timeout_rate < 15 else ("Critical" if timeout_rate > 25 else "Warning"), "", ""),
            ("A", "A1", "Overall_Accuracy", "Độ chính xác trung bình (parse transaction)", "%", "Test results", "≥ 80%", "< 70%", f"{avg_accuracy:.1f}", "OK" if avg_accuracy >= 80 else ("Critical" if avg_accuracy < 70 else "Warning"), "LLM04,LLM09", "Scaffolding"),
            ("A", "A2", "Pass_Rate", "Tỷ lệ test case Pass", "%", "Test results", "≥ 80%", "< 70%", f"{summary.pass_rate():.1f}", "OK" if summary.pass_rate() >= 80 else ("Critical" if summary.pass_rate() < 70 else "Warning"), "LLM04,LLM09", "Step-by-step"),
            ("A", "A3", "Partial_Rate", "Tỷ lệ test case Partial Pass", "%", "Test results", "", "", f"{(summary.partial / max(summary.total_tests, 1) * 100):.1f}", "Info", "", ""),
            ("S", "S1", "Total_Tests_Executed", "Số test case đã chạy", "count", "Test runner", "", "", str(summary.total_tests), "Info", "", ""),
            ("S", "S2", "Tests_Per_Minute", "Tốc độ chạy test", "tests/min", "Calculated", "", "", f"{len(results) / max(summary.duration_seconds() / 60, 1):.1f}" if hasattr(summary, 'duration_seconds') else "N/A", "Info", "", ""),
            ("Stability", "ST1", "Error_Rate", "Tỷ lệ request lỗi (crash, exception)", "%", "Test results", "< 1%", "> 3%", f"{error_rate:.1f}", "OK" if error_rate < 1 else ("Critical" if error_rate > 3 else "Warning"), "", ""),
            ("Stability", "ST2", "Stability_Issues", "Số lượng stability issues (high latency, error)", "count", "Test results", "≤ 10", "> 20", str(summary.stability_issues), "OK" if summary.stability_issues <= 10 else ("Critical" if summary.stability_issues > 20 else "Warning"), "", ""),
            ("Stability", "ST3", "Stability_Rate", "Tỷ lệ request ổn định (không error)", "%", "Test results", "≥ 95%", "< 90%", f"{100 - error_rate:.1f}", "OK" if (100 - error_rate) >= 95 else "Warning", "", ""),
            ("Security", "SEC1", "Security_Issues", "Số lượng security issues phát hiện", "count", "Test results", "0", "> 0", str(summary.security_issues), "OK" if summary.security_issues == 0 else "Critical", "LLM01-LLM10", ""),
            ("Security", "SEC2", "Security_Pass_Rate", "Tỷ lệ security tests Pass", "%", "Test results", "100%", "< 95%", f"{self._calc_security_pass_rate(results):.1f}", "OK" if self._calc_security_pass_rate(results) >= 100 else "Warning", "LLM01-LLM10", ""),
        ]
        
        for row_idx, row_data in enumerate(metrics_data, start=2):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border
                # Color status column
                if col == 10:
                    if value == "OK":
                        cell.fill = self.pass_fill
                    elif value == "Warning":
                        cell.fill = self.partial_fill
                    elif value == "Critical":
                        cell.fill = self.fail_fill
        
        # Set column widths
        widths = [12, 10, 25, 55, 18, 20, 18, 15, 15, 10, 20, 25]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w
    
    def _create_owasp_coverage(self, ws, results: List[TestRunResult]):
        """Create 04_OWASP_Coverage sheet"""
        headers = [
            "OWASP_ID", "Risk_Name", "Relevant_to_System", "Mitigation_Summary",
            "Owner", "Status", "Last_Reviewed_Date", "Related_Test_Cases", "Pass_Rate_%"
        ]
        self._apply_header_style(ws, 1, headers)
        
        # Calculate coverage per OWASP risk
        owasp_stats = {}
        for risk_id in OWASP_RISKS.keys():
            owasp_stats[risk_id] = {"test_cases": [], "passed": 0, "total": 0}
        
        for result in results:
            if result.owasp_check:
                for risk_id, status in result.owasp_check.items():
                    if risk_id in owasp_stats:
                        owasp_stats[risk_id]["test_cases"].append(result.test_case_id)
                        owasp_stats[risk_id]["total"] += 1
                        if status == "OK":
                            owasp_stats[risk_id]["passed"] += 1
        
        row_idx = 2
        for risk_id, risk_info in OWASP_RISKS.items():
            stats = owasp_stats.get(risk_id, {"test_cases": [], "passed": 0, "total": 0})
            pass_rate = (stats["passed"] / stats["total"] * 100) if stats["total"] > 0 else 0
            
            ws.cell(row=row_idx, column=1, value=risk_id).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=risk_info["name"]).border = self.thin_border
            ws.cell(row=row_idx, column=3, value="Yes" if stats["total"] > 0 else "Not tested").border = self.thin_border
            ws.cell(row=row_idx, column=4, value=risk_info.get("mitigation", "")).border = self.thin_border
            ws.cell(row=row_idx, column=5, value="Security Team").border = self.thin_border
            
            status = "Tested" if stats["total"] > 0 else "Not started"
            status_cell = ws.cell(row=row_idx, column=6, value=status)
            status_cell.border = self.thin_border
            
            ws.cell(row=row_idx, column=7, value=datetime.now().strftime("%Y-%m-%d") if stats["total"] > 0 else "").border = self.thin_border
            ws.cell(row=row_idx, column=8, value=", ".join(stats["test_cases"][:5]) + ("..." if len(stats["test_cases"]) > 5 else "")).border = self.thin_border
            
            rate_cell = ws.cell(row=row_idx, column=9, value=f"{pass_rate:.1f}" if stats["total"] > 0 else "N/A")
            rate_cell.border = self.thin_border
            if stats["total"] > 0:
                if pass_rate >= 80:
                    rate_cell.fill = self.pass_fill
                elif pass_rate >= 50:
                    rate_cell.fill = self.partial_fill
                else:
                    rate_cell.fill = self.fail_fill
            
            row_idx += 1
        
        # Set column widths
        widths = [10, 35, 18, 40, 15, 15, 18, 40, 12]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w

    def _create_owasp_test_results(self, ws, results: List[TestRunResult]):
        """Create 05_OWASP_Test_Results sheet"""
        headers = [
            "OWASP_Test_ID", "OWASP_ID", "Test_Case_ID", "Scenario_Description",
            "User_Message_Input", "Expected_Behavior", "Actual_Behavior",
            "Pass_Fail", "Severity_if_Failed", "Notes"
        ]
        self._apply_header_style(ws, 1, headers)
        
        row_idx = 2
        owasp_test_counter = 1
        
        for result in results:
            if result.owasp_check:
                for risk_id, status in result.owasp_check.items():
                    ws.cell(row=row_idx, column=1, value=f"OWASP_TC_{owasp_test_counter:03d}").border = self.thin_border
                    ws.cell(row=row_idx, column=2, value=risk_id).border = self.thin_border
                    ws.cell(row=row_idx, column=3, value=result.test_case_id).border = self.thin_border
                    ws.cell(row=row_idx, column=4, value=OWASP_RISKS.get(risk_id, {}).get("name", "")).border = self.thin_border
                    
                    # User message (truncated)
                    user_msg = result.actual_bot_response[:100] + "..." if result.actual_bot_response and len(result.actual_bot_response) > 100 else ""
                    ws.cell(row=row_idx, column=5, value=user_msg).border = self.thin_border
                    
                    ws.cell(row=row_idx, column=6, value=f"Bot should handle {risk_id} securely").border = self.thin_border
                    ws.cell(row=row_idx, column=7, value=status).border = self.thin_border
                    
                    # Pass/Fail
                    pass_fail = "Pass" if status == "OK" else "Fail"
                    pf_cell = ws.cell(row=row_idx, column=8, value=pass_fail)
                    pf_cell.border = self.thin_border
                    if pass_fail == "Pass":
                        pf_cell.fill = self.pass_fill
                    else:
                        pf_cell.fill = self.fail_fill
                    
                    # Severity
                    severity = OWASP_RISKS.get(risk_id, {}).get("severity", "Medium") if pass_fail == "Fail" else ""
                    ws.cell(row=row_idx, column=9, value=severity).border = self.thin_border
                    ws.cell(row=row_idx, column=10, value=result.notes or "").border = self.thin_border
                    
                    row_idx += 1
                    owasp_test_counter += 1
        
        # Set column widths
        widths = [15, 10, 15, 35, 50, 40, 30, 10, 18, 40]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w
    
    def _create_class_checklist(self, ws, results: List[TestRunResult]):
        """Create 06_CLASS_Checklist sheet"""
        headers = [
            "CLASS_Component", "Description", "Implemented_YN", "Evidence_Link",
            "Test_Cases", "Pass_Rate_%", "Notes"
        ]
        self._apply_header_style(ws, 1, headers)
        
        # Calculate stats per CLASS principle
        class_stats = {}
        for principle in CLASS_PRINCIPLES.keys():
            class_stats[principle] = {"test_cases": [], "passed": 0, "total": 0}
        
        for result in results:
            if result.class_principles_check:
                for principle, passed in result.class_principles_check.items():
                    if principle in class_stats:
                        class_stats[principle]["test_cases"].append(result.test_case_id)
                        class_stats[principle]["total"] += 1
                        if passed:
                            class_stats[principle]["passed"] += 1
        
        row_idx = 2
        for principle, principle_info in CLASS_PRINCIPLES.items():
            stats = class_stats.get(principle, {"test_cases": [], "passed": 0, "total": 0})
            pass_rate = (stats["passed"] / stats["total"] * 100) if stats["total"] > 0 else 0
            
            # Get description - handle both dict and string
            if isinstance(principle_info, dict):
                description = principle_info.get("description", "")
            else:
                description = str(principle_info)
            
            ws.cell(row=row_idx, column=1, value=principle).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=description).border = self.thin_border
            
            # Implemented status
            impl_status = "Yes" if stats["total"] > 0 and pass_rate >= 80 else ("Partial" if stats["total"] > 0 else "No")
            impl_cell = ws.cell(row=row_idx, column=3, value=impl_status)
            impl_cell.border = self.thin_border
            if impl_status == "Yes":
                impl_cell.fill = self.pass_fill
            elif impl_status == "Partial":
                impl_cell.fill = self.partial_fill
            else:
                impl_cell.fill = self.fail_fill
            
            ws.cell(row=row_idx, column=4, value="").border = self.thin_border
            ws.cell(row=row_idx, column=5, value=", ".join(stats["test_cases"][:5]) + ("..." if len(stats["test_cases"]) > 5 else "")).border = self.thin_border
            
            rate_cell = ws.cell(row=row_idx, column=6, value=f"{pass_rate:.1f}" if stats["total"] > 0 else "N/A")
            rate_cell.border = self.thin_border
            
            ws.cell(row=row_idx, column=7, value="").border = self.thin_border
            
            row_idx += 1
        
        # Set column widths
        widths = [30, 70, 15, 20, 40, 12, 40]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w
    def _create_class_metrics_explanation(self, ws):
        """Create 05_CLASS_Metrics_Explanation sheet"""
        headers = ["Dimension", "Metric", "Description", "Unit", "Threshold", "Calculation", "Notes"]
        self._apply_header_style(ws, 1, headers)
        
        data = [
            # C - Cost
            ("C", "Cost", "Chi phí vận hành chatbot", "VND", "< 1000 (simple), < 5000 (complex)", 
             "cost_vnd = (input_tokens/1000 * input_rate + output_tokens/1000 * output_rate) * usd_to_vnd_rate",
             "Chi phí phụ thuộc vào token usage và model pricing"),
            ("C", "Prompt Tokens", "Số token trong input prompt", "tokens", "< 500 (simple), < 2000 (complex)",
             "Đếm từ prompt gửi đến LLM", "Prompt càng dài, cost càng cao"),
            ("C", "Completion Tokens", "Số token trong output response", "tokens", "< 200 (simple), < 800 (complex)",
             "Đếm từ response nhận về", "Response càng dài, cost càng cao"),
            ("C", "Total Tokens", "Tổng số token", "tokens", "< 700 (simple), < 2800 (complex)",
             "prompt_tokens + completion_tokens", "Tổng token usage"),
            
            # L - Latency
            ("L", "Latency", "Thời gian phản hồi", "ms", "< 3000 (simple), < 5000 (complex)",
             "end_time - start_time", "Thời gian từ request đến response"),
            ("L", "P50 Latency", "Median latency (50th percentile)", "ms", "< 2000",
             "50th percentile của latencies", "50% requests < P50"),
            ("L", "P95 Latency", "95th percentile latency", "ms", "< 5000",
             "95th percentile của latencies", "95% requests < P95"),
            ("L", "P99 Latency", "99th percentile latency", "ms", "< 8000",
             "99th percentile của latencies", "99% requests < P99"),
            
            # A - Accuracy
            ("A", "Accuracy Score", "Độ chính xác tổng thể", "%", ">= 80",
             "(correct_fields / total_fields) * 100", "Tỷ lệ fields parse đúng"),
            ("A", "Amount Accuracy", "Độ chính xác parse amount", "%", ">= 95",
             "amount_correct / amount_total * 100", "Critical field, phải >= 95%"),
            ("A", "Category Accuracy", "Độ chính xác parse category", "%", ">= 85",
             "category_correct / category_total * 100", "Medium importance"),
            ("A", "Type Accuracy", "Độ chính xác detect transaction type", "%", ">= 90",
             "type_correct / type_total * 100", "High importance"),
            ("A", "Date Accuracy", "Độ chính xác parse date", "%", ">= 80",
             "date_correct / date_total * 100", "Low importance, >= 80% acceptable"),
            ("A", "Intent Accuracy", "Độ chính xác detect intent", "%", ">= 90",
             "intent_correct / intent_total * 100", "High importance"),
            
            # S1 - Scalability
            ("S1", "Concurrent Users", "Số users đồng thời", "count", "10, 50, 100, 500",
             "Số requests đồng thời", "Test với các mức khác nhau"),
            ("S1", "Throughput", "Requests per second", "rps", ">= 50",
             "total_requests / duration_seconds", "Số requests xử lý được mỗi giây"),
            ("S1", "Success Rate", "Tỷ lệ thành công", "%", ">= 95",
             "successful_requests / total_requests * 100", "Phải >= 95%"),
            ("S1", "Error Rate", "Tỷ lệ lỗi", "%", "< 1",
             "failed_requests / total_requests * 100", "Phải < 1%"),
            
            # S2 - Stability
            ("S2", "No Crash", "Không crash", "boolean", "True",
             "Check if system crashed", "System phải không crash"),
            ("S2", "Meaningful Response", "Response có ý nghĩa", "boolean", "True",
             "Check if response is meaningful", "Response phải có ý nghĩa"),
            ("S2", "Consistency", "Tính nhất quán", "boolean", "True",
             "Check if same input gives same output", "Cùng input phải cho cùng output"),
            ("S2", "Error Handling", "Xử lý lỗi đúng cách", "boolean", "True",
             "Check if errors are handled gracefully", "Lỗi phải được handle đúng cách"),
        ]
        
        row_idx = 2
        for row_data in data:
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.thin_border
                if col == 1:  # Dimension column
                    cell.font = Font(bold=True)
            row_idx += 1
        
        # Set column widths
        widths = [12, 20, 40, 10, 30, 60, 50]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w
    
    def _create_thresholds_comparison(self, ws, results: List[TestRunResult], summary: TestSummary):
        """Create 06_Thresholds_Comparison sheet"""
        headers = [
            "Metric", "Threshold", "Actual", "Status", "Difference", "Percentage", "Notes"
        ]
        self._apply_header_style(ws, 1, headers)
        
        # Calculate actual values
        latencies = [r.measured_latency_ms for r in results if r.measured_latency_ms > 0]
        costs = [r.measured_cost_vnd for r in results if r.measured_cost_vnd > 0]
        accuracies = [r.accuracy_score_percent for r in results if r.accuracy_score_percent > 0]
        
        avg_latency = sum(latencies) / len(latencies) if latencies else 0
        p95_latency = sorted(latencies)[int(len(latencies) * 0.95)] if len(latencies) > 0 else 0
        avg_cost = sum(costs) / len(costs) if costs else 0
        avg_accuracy = sum(accuracies) / len(accuracies) if accuracies else 0
        
        # Compare with thresholds
        comparisons = [
            {
                "metric": "Average Latency",
                "threshold": self.config.latency_warning_ms,
                "actual": avg_latency,
                "unit": "ms"
            },
            {
                "metric": "P95 Latency",
                "threshold": self.config.p95_latency_max_ms,
                "actual": p95_latency,
                "unit": "ms"
            },
            {
                "metric": "Average Cost per Request",
                "threshold": self.config.cost_per_request_max_vnd_simple,
                "actual": avg_cost,
                "unit": "VND"
            },
            {
                "metric": "Average Accuracy",
                "threshold": self.config.accuracy_pass_threshold * 100,
                "actual": avg_accuracy,
                "unit": "%"
            },
            {
                "metric": "Success Rate",
                "threshold": self.config.success_rate_min * 100,
                "actual": summary.pass_rate(),
                "unit": "%"
            },
        ]
        
        row_idx = 2
        for comp in comparisons:
            threshold = comp["threshold"]
            actual = comp["actual"]
            unit = comp["unit"]
            
            # Determine status
            if comp["metric"] in ["Average Latency", "P95 Latency", "Average Cost per Request"]:
                # Lower is better
                if actual <= threshold:
                    status = "✅ Pass"
                    status_color = self.pass_fill
                elif actual <= threshold * 1.5:
                    status = "⚠️ Warning"
                    status_color = self.partial_fill
                else:
                    status = "❌ Fail"
                    status_color = self.fail_fill
            else:
                # Higher is better
                if actual >= threshold:
                    status = "✅ Pass"
                    status_color = self.pass_fill
                elif actual >= threshold * 0.9:
                    status = "⚠️ Warning"
                    status_color = self.partial_fill
                else:
                    status = "❌ Fail"
                    status_color = self.fail_fill
            
            # Calculate difference
            if comp["metric"] in ["Average Latency", "P95 Latency", "Average Cost per Request"]:
                diff = actual - threshold
                percentage = (actual / threshold - 1) * 100 if threshold > 0 else 0
            else:
                diff = actual - threshold
                percentage = (actual / threshold - 1) * 100 if threshold > 0 else 0
            
            ws.cell(row=row_idx, column=1, value=comp["metric"]).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=f"{threshold} {unit}").border = self.thin_border
            ws.cell(row=row_idx, column=3, value=f"{actual:.2f} {unit}").border = self.thin_border
            
            status_cell = ws.cell(row=row_idx, column=4, value=status)
            status_cell.border = self.thin_border
            status_cell.fill = status_color
            
            ws.cell(row=row_idx, column=5, value=f"{diff:+.2f} {unit}").border = self.thin_border
            ws.cell(row=row_idx, column=6, value=f"{percentage:+.1f}%").border = self.thin_border
            
            # Notes
            if status == "❌ Fail":
                notes = f"Actual value {'exceeds' if diff > 0 else 'below'} threshold by {abs(percentage):.1f}%"
            elif status == "⚠️ Warning":
                notes = f"Approaching threshold limit"
            else:
                notes = "Within acceptable range"
            
            ws.cell(row=row_idx, column=7, value=notes).border = self.thin_border
            
            row_idx += 1
        
        # Set column widths
        widths = [25, 20, 20, 15, 20, 15, 40]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w
    
    def _create_workload_analysis(self, ws, results: List[TestRunResult], summary: TestSummary):
        """Create 07_Workload_Analysis sheet"""
        headers = [
            "Workload_Level", "Concurrent_Users", "Total_Requests", "Successful", "Failed",
            "Success_Rate_%", "Error_Rate_%", "Avg_Latency_ms", "P95_Latency_ms", "P99_Latency_ms",
            "Throughput_rps", "Avg_Cost_VND", "Status", "Notes"
        ]
        self._apply_header_style(ws, 1, headers)
        
        # Group results by workload level (if available)
        # For now, show overall statistics
        latencies = [r.measured_latency_ms for r in results if r.measured_latency_ms > 0]
        costs = [r.measured_cost_vnd for r in results if r.measured_cost_vnd > 0]
        
        if latencies:
            sorted_latencies = sorted(latencies)
            p50 = sorted_latencies[int(len(sorted_latencies) * 0.50)] if len(sorted_latencies) > 0 else 0
            p95 = sorted_latencies[int(len(sorted_latencies) * 0.95)] if len(sorted_latencies) > 0 else 0
            p99 = sorted_latencies[int(len(sorted_latencies) * 0.99)] if len(sorted_latencies) > 0 else 0
        else:
            p50 = p95 = p99 = 0
        
        avg_latency = sum(latencies) / len(latencies) if latencies else 0
        avg_cost = sum(costs) / len(costs) if costs else 0
        
        # Calculate success/error rates
        total = summary.total_tests
        successful = summary.passed
        failed = summary.failed
        success_rate = (successful / total * 100) if total > 0 else 0
        error_rate = (failed / total * 100) if total > 0 else 0
        
        # Estimate throughput (assuming test duration)
        # This is a placeholder - actual throughput should be measured during concurrent tests
        estimated_throughput = 0  # Would need actual test duration
        
        # Determine status
        if success_rate >= 95 and error_rate < 1 and p95 < self.config.p95_latency_max_ms:
            status = "✅ Pass"
            status_color = self.pass_fill
        elif success_rate >= 90 and error_rate < 5:
            status = "⚠️ Warning"
            status_color = self.partial_fill
        else:
            status = "❌ Fail"
            status_color = self.fail_fill
        
        row_idx = 2
        # Overall row
        ws.cell(row=row_idx, column=1, value="Overall").border = self.thin_border
        ws.cell(row=row_idx, column=2, value="N/A").border = self.thin_border
        ws.cell(row=row_idx, column=3, value=total).border = self.thin_border
        ws.cell(row=row_idx, column=4, value=successful).border = self.thin_border
        ws.cell(row=row_idx, column=5, value=failed).border = self.thin_border
        ws.cell(row=row_idx, column=6, value=f"{success_rate:.1f}").border = self.thin_border
        ws.cell(row=row_idx, column=7, value=f"{error_rate:.1f}").border = self.thin_border
        ws.cell(row=row_idx, column=8, value=f"{avg_latency:.0f}").border = self.thin_border
        ws.cell(row=row_idx, column=9, value=f"{p95:.0f}").border = self.thin_border
        ws.cell(row=row_idx, column=10, value=f"{p99:.0f}").border = self.thin_border
        ws.cell(row=row_idx, column=11, value=f"{estimated_throughput:.1f}").border = self.thin_border
        ws.cell(row=row_idx, column=12, value=f"{avg_cost:.0f}").border = self.thin_border
        
        status_cell = ws.cell(row=row_idx, column=13, value=status)
        status_cell.border = self.thin_border
        status_cell.fill = status_color
        
        notes = f"Overall test results. P50 latency: {p50:.0f}ms"
        ws.cell(row=row_idx, column=14, value=notes).border = self.thin_border
        
        # Set column widths
        widths = [18, 15, 12, 12, 12, 12, 12, 15, 15, 15, 15, 15, 12, 40]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w

